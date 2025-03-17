# modules/report.py

import os
import pandas as pd
from colorama import init, Fore, Style

# Initialize colorama
init()


def output_message(symbol, message, level="info"):
    """Generic function to print messages with different colors and symbols based on the level."""
    colors = {
        "good": Fore.GREEN + Style.BRIGHT,
        "warning": Fore.YELLOW + Style.BRIGHT,
        "bad": Fore.RED + Style.BRIGHT,
        "indifferent": Fore.BLUE + Style.BRIGHT,
        "error": Fore.RED + Style.BRIGHT + "!!! ",
        "info": Fore.WHITE + Style.BRIGHT,
    }
    color = colors.get(level, Fore.WHITE + Style.BRIGHT)
    print(color + f"{symbol} {message}" + Style.RESET_ALL)


def write_to_excel(data, file_name="output.xlsx"):
    """Write data to an Excel file with enhanced formatting and color coding."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule

    # Create a new workbook and active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spoofy Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Alignment for all cells
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define fills for spoofing status
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for spoofable
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow for maybe
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for not spoofable
    
    # Get all keys from the first data entry
    if not data:
        wb.save(file_name)
        return
    
    # Reorder columns to place important ones first
    important_fields = [
        "DOMAIN", "SPOOFING_POSSIBLE", "SPOOFING_TYPE", "DOMAIN_TYPE",
        "SPF", "SPF_MULTIPLE_ALLS", "SPF_NUM_DNS_QUERIES", "SPF_TOO_MANY_DNS_QUERIES",
        "DMARC", "DMARC_POLICY", "DMARC_PCT", "DMARC_ASPF", "DMARC_SP",
        "DMARC_FORENSIC_REPORT", "DMARC_AGGREGATE_REPORT",
        "BIMI_RECORD", "BIMI_VERSION", "BIMI_LOCATION", "BIMI_AUTHORITY",
        "DNS_SERVER"
    ]
    
    # Ensure all keys from data are included (in case there are new keys not in our ordering)
    all_keys = set()
    for entry in data:
        all_keys.update(entry.keys())
    
    headers = [field for field in important_fields if field in all_keys]
    headers.extend([field for field in all_keys if field not in important_fields])
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Write data
    for row_num, entry in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = entry.get(header, "")
            cell.alignment = left_alignment if col_num > 2 else center_alignment
            cell.border = thin_border
            
            # Color the entire row based on spoofing status
            if header == "SPOOFING_POSSIBLE":
                spoofing_possible = entry.get(header)
                if spoofing_possible is True:
                    cell.value = "YES"
                    cell.font = Font(color="9C0006", bold=True)
                elif spoofing_possible is False:
                    cell.value = "NO"
                    cell.font = Font(color="006100", bold=True)
                else:  # None or maybe
                    cell.value = "MAYBE"
                    cell.font = Font(color="9C5700", bold=True)
            
            # Set coloring based on the spoofing status
            spoofing_possible = entry.get("SPOOFING_POSSIBLE")
            if spoofing_possible is True:
                cell.fill = red_fill
            elif spoofing_possible is False:
                cell.fill = green_fill
            elif spoofing_possible is None:  # maybe
                cell.fill = yellow_fill
    
    # Auto-adjust column width based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Freeze the top row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(file_name)
    return


def printer(**kwargs):
    """Utility function to print the results of DMARC, SPF, and BIMI checks in the original format."""
    domain = kwargs.get("DOMAIN")
    subdomain = kwargs.get("DOMAIN_TYPE") == "subdomain"
    dns_server = kwargs.get("DNS_SERVER")
    spf_record = kwargs.get("SPF")
    spf_all = kwargs.get("SPF_MULTIPLE_ALLS")
    spf_dns_query_count = kwargs.get("SPF_NUM_DNS_QUERIES")
    dmarc_record = kwargs.get("DMARC")
    p = kwargs.get("DMARC_POLICY")
    pct = kwargs.get("DMARC_PCT")
    aspf = kwargs.get("DMARC_ASPF")
    sp = kwargs.get("DMARC_SP")
    fo = kwargs.get("DMARC_FORENSIC_REPORT")
    rua = kwargs.get("DMARC_AGGREGATE_REPORT")
    bimi_record = kwargs.get("BIMI_RECORD")
    vbimi = kwargs.get("BIMI_VERSION")
    location = kwargs.get("BIMI_LOCATION")
    authority = kwargs.get("BIMI_AUTHORITY")
    spoofable = kwargs.get("SPOOFING_POSSIBLE")
    spoofing_type = kwargs.get("SPOOFING_TYPE")

    output_message("[*]", f"Domain: {domain}", "indifferent")
    output_message("[*]", f"Is subdomain: {subdomain}", "indifferent")
    output_message("[*]", f"DNS Server: {dns_server}", "indifferent")

    if spf_record:
        output_message("[*]", f"SPF record: {spf_record}", "info")
        if spf_all is None:
            output_message("[*]", "SPF does not contain an `All` item.", "info")
        elif spf_all == "2many":
            output_message(
                "[?]", "SPF record contains multiple `All` items.", "warning"
            )
        else:
            output_message("[*]", f"SPF all record: {spf_all}", "info")
        output_message(
            "[*]",
            f"SPF DNS query count: {spf_dns_query_count}"
            if spf_dns_query_count <= 10
            else f"Too many SPF DNS query lookups {spf_dns_query_count}.",
            "info",
        )
    else:
        output_message("[?]", "No SPF record found.", "warning")

    if dmarc_record:
        output_message("[*]", f"DMARC record: {dmarc_record}", "info")
        output_message(
            "[*]", f"Found DMARC policy: {p}" if p else "No DMARC policy found.", "info"
        )
        output_message(
            "[*]", f"Found DMARC pct: {pct}" if pct else "No DMARC pct found.", "info"
        )
        output_message(
            "[*]",
            f"Found DMARC aspf: {aspf}" if aspf else "No DMARC aspf found.",
            "info",
        )
        output_message(
            "[*]",
            f"Found DMARC subdomain policy: {sp}"
            if sp
            else "No DMARC subdomain policy found.",
            "info",
        )
        output_message(
            "[*]",
            f"Forensics reports will be sent: {fo}"
            if fo
            else "No DMARC forensics report location found.",
            "indifferent",
        )
        output_message(
            "[*]",
            f"Aggregate reports will be sent to: {rua}"
            if rua
            else "No DMARC aggregate report location found.",
            "indifferent",
        )
    else:
        output_message("[?]", "No DMARC record found.", "warning")

    if bimi_record:
        output_message("[*]", f"BIMI record: {bimi_record}", "info")
        output_message("[*]", f"BIMI version: {vbimi}", "info")
        output_message("[*]", f"BIMI location: {location}", "info")
        output_message("[*]", f"BIMI authority: {authority}", "info")

    if spoofing_type:
        level = "good" if spoofable else "bad"
        symbol = "[+]" if level == "good" else "[-]"
        output_message(symbol, spoofing_type, level)

    print()  # Padding
